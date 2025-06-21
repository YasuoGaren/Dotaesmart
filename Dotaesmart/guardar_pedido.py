from flask import Flask, request, jsonify, render_template
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

app = Flask(__name__,
            template_folder='.',  
            static_folder='.')  
                                 


CORS(app)

CARPETA_PEDIDOS = "pedidos" 
ARCHIVO_EXCEL_NOMBRE = "pedidos.xlsx"
ARCHIVO_EXCEL_RUTA_COMPLETA = os.path.join(CARPETA_PEDIDOS, ARCHIVO_EXCEL_NOMBRE)

@app.route('/')
def index():
    try:
        return render_template('test.html')
    except Exception as e:
        print(f"Error al renderizar la plantilla 'dotasmesmart.html': {e}")
        return f"<h1>Error al cargar la página principal</h1><p>Detalle: {e}</p><p>Verifica si 'dotasmesmart.html' está en la carpeta correcta y sin errores.</p>", 500


@app.route("/guardar_pedido", methods=["POST"])
def guardar_pedido():
    try:
        datos = request.json
        nombre = datos.get("nombre")
        direccion = datos.get("direccion")
        contacto = datos.get("contacto")
        carrito = datos.get("carrito", [])

        os.makedirs(CARPETA_PEDIDOS, exist_ok=True)


        # Verifica si el archivo Excel existe.
        if not os.path.exists(ARCHIVO_EXCEL_RUTA_COMPLETA):
            wb = Workbook()
            ws = wb.active
            ws.title = "Pedidos"
            ws.append(["Fecha", "Nombre", "Dirección", "Contacto", "Producto", "Cantidad", "Precio Unitario", "Subtotal"])
            wb.save(ARCHIVO_EXCEL_RUTA_COMPLETA)

        wb = load_workbook(ARCHIVO_EXCEL_RUTA_COMPLETA)
        ws = wb.active

        for item in carrito:
            nombre_producto = item["nombre"]
            cantidad = item["cantidad"]
            precio = item["precio"]
            subtotal = cantidad * precio
            ws.append([
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                nombre,
                direccion,
                contacto,
                nombre_producto,
                cantidad,
                precio,
                subtotal
            ])

        wb.save(ARCHIVO_EXCEL_RUTA_COMPLETA)
        return jsonify({"mensaje": "Pedido guardado con éxito"}), 200

    except Exception as e:
        print("Error al guardar el pedido:", e)
        return jsonify({"error": "Error interno del servidor", "detalle": str(e)}), 500
